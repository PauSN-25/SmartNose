import pandas as pd
import numpy as np
import os
import re
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches 
import seaborn as sns
import warnings
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Configuración
warnings.filterwarnings("ignore")
plt.rcParams.update({'figure.max_open_warning': 0})


# 1. FUNCIONES DE UTILIDAD Y ESTILO 

def clave_orden_natural(texto):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(texto))]

# --- FUNCIÓN PARA ASIGNAR COLORES A LOS PARÁMETROS ---
def asignar_color_parametro(nombre_variable):
    nombre = str(nombre_variable).lower()
    if 'area' in nombre:
        return '#2ca02c' # Verde
    elif 't_max' in nombre:
        return '#1f77b4' # Azul
    elif 'max' in nombre: 
        return '#ff7f0e' # Naranja
    elif 'slope' in nombre:
        return '#9467bd' # Morado
    else:
        return '#7f7f7f' # Gris
# -------------------------------------------------------------------

def leer_datos_limpios(ruta_archivo):
    try:
        df = pd.read_csv(ruta_archivo, comment='#')
        df.columns = [c.strip() for c in df.columns]
        return df
    except Exception as e:
        print(f"⚠️ Error leyendo {os.path.basename(ruta_archivo)}: {e}")
        return None

def auto_ajustar_columnas(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        try:
            col_letter = column_cells[0].column_letter
        except AttributeError:
            col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 3

def aplicar_colores_rsd(ws):
    fill_verde   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amarillo= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    fill_rojo    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    col_idx = None
    for cell in ws[1]: 
        if cell.value == "RSD_Global":
            col_idx = cell.column
            break
    
    if col_idx is None: return

    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        valor = cell.value
        if valor is None or not isinstance(valor, (int, float)): continue

        if valor < 10.0: cell.fill = fill_verde
        elif 10.0 <= valor < 20.0: cell.fill = fill_amarillo
        elif 20.0 <= valor < 30.0: cell.fill = fill_naranja
        else: cell.fill = fill_rojo

def aplicar_colores_pearson(ws):
    fill_verde   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amarillo= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    fill_rojo    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    cols_interes = ["Pearson_Global", "Pearson_Min"]
    indices = []

    for cell in ws[1]:
        if cell.value in cols_interes:
            indices.append(cell.column)
    
    if not indices: return

    for col_idx in indices:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            valor = cell.value
            if valor is None or not isinstance(valor, (int, float)): continue

            if valor >= 0.90: cell.fill = fill_verde
            elif valor >= 0.80: cell.fill = fill_amarillo
            elif valor >= 0.70: cell.fill = fill_naranja
            else: cell.fill = fill_rojo


# 2. METADATOS


def extraer_metadatos_csv(ruta_archivo):
    
    meta = {
        "CÓDIGO": "", "MUESTRA": "", "TIEMPO": 0, "FECHA": "", "USUARIO": "",
        "ARCHIVO ORIGEN": os.path.basename(ruta_archivo), "CARPETA": ""
    }
    duraciones = []
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            for _ in range(30):  # Leemos hasta 30 líneas buscando metadatos
                linea = f.readline().strip()
                
                # Si llegamos a los datos reales o a la cabecera (sin #), paramos de buscar
                if not linea.startswith("#"): break
                
                # Extracción flexible
                if "# Nombre Ensayo:" in linea: 
                    meta["CÓDIGO"] = linea.split(":", 1)[1].strip()
                elif "# Muestra:" in linea: 
                    meta["MUESTRA"] = linea.split(":", 1)[1].strip()
                elif "# Fecha:" in linea or "# Fecha y Hora de creacion archivo:" in linea: 
                    
                    meta["FECHA"] = linea.split(":", 1)[1].strip()
                elif "# Usuario:" in linea: 
                    meta["USUARIO"] = linea.split(":", 1)[1].strip()
                elif "# Carpeta:" in linea: 
                    meta["CARPETA"] = linea.split(":", 1)[1].strip()
                elif "# Duracion (min):" in linea: 
                    duraciones.append(linea.split(":", 1)[1].strip())

        # El tiempo de medición es la segunda duración listada (Fase 1)
        if len(duraciones) >= 2: meta["TIEMPO"] = duraciones[1]
        elif len(duraciones) == 1: meta["TIEMPO"] = duraciones[0]
            
    except Exception as e: 
        print(f"⚠️ Error leyendo metadatos de {os.path.basename(ruta_archivo)}: {e}")
        
    return meta

# 3. PROCESAMIENTO CIENTÍFICO

def calcular_normalizacion_sensor(df, columna_sensor):
    datos_vent = df[df.iloc[:, 0].isin(["VENT", "VENTILACION"])][columna_sensor]
    datos_vent_reales = datos_vent.replace(0, np.nan).dropna()
    
    if len(datos_vent_reales) < 1: return None
    baseline = datos_vent_reales.tail(5).mean()
    if baseline == 0: return None

    datos_med = df[df.iloc[:, 0].isin(["MED", "MEDICION"])].copy().reset_index(drop=True)
    valores_med = datos_med[columna_sensor]
    
    datos_norm = ((baseline - valores_med) / baseline) * (-100.0) + 100.0
    datos_norm = datos_norm.replace(0, np.nan)
    return datos_norm

def obtener_curvas_normalizadas_archivo(ruta_archivo):
    df = leer_datos_limpios(ruta_archivo)
    if df is None: return None
    
    sensores_gas = [f"{L}{i}" for L in ['SA', 'SB', 'SC'] for i in range(1, 11)]
    curvas = {}
    
    for sensor in sensores_gas:
        if sensor in df.columns:
            serie_norm = calcular_normalizacion_sensor(df, sensor)
            if serie_norm is not None:
                curvas[sensor] = serie_norm
                
    if not curvas: return None
    
    df_curvas = pd.DataFrame(curvas)
    df_med = df[df.iloc[:, 0].isin(["MED", "MEDICION"])].reset_index(drop=True)
    cols_tiempo = [c for c in df_med.columns if 'tiempo' in c.lower()]
    if cols_tiempo:
        df_curvas['Tiempo'] = df_med[cols_tiempo[0]]
    else:
        df_curvas['Tiempo'] = range(len(df_curvas))
        
    return df_curvas

def procesar_grupo_ensayos(lista_archivos):
    acumulado_gases = {}
    acumulado_ambientales = {'Temp': [], 'Hum': [], 'Pres': []}
    sensores_gas = [f"{L}{i}" for L in ['SA', 'SB', 'SC'] for i in range(1, 11)]

    df_referencia_tiempo = None

    for archivo in lista_archivos:
        df = leer_datos_limpios(archivo)
        if df is None: continue
        
        df_med = df[df.iloc[:, 0].isin(["MED", "MEDICION"])].reset_index(drop=True)
        if df_referencia_tiempo is None:
            cols_tiempo = [c for c in df_med.columns if 'tiempo' in c.lower()]
            if cols_tiempo: df_referencia_tiempo = df_med[cols_tiempo[0]]
            else: df_referencia_tiempo = pd.Series(range(len(df_med)), name="Tiempo")

        for var in ['Temp', 'Hum', 'Pres']:
            cols = [c for c in df_med.columns if var in c]
            if cols: acumulado_ambientales[var].append(df_med[cols].mean(axis=1))

        for sensor in sensores_gas:
            if sensor in df.columns:
                serie_norm = calcular_normalizacion_sensor(df, sensor)
                if serie_norm is not None:
                    if sensor not in acumulado_gases: acumulado_gases[sensor] = []
                    acumulado_gases[sensor].append(serie_norm)

    df_final = pd.DataFrame()
    if df_referencia_tiempo is not None: df_final['Tiempo'] = df_referencia_tiempo

    for var in ['Temp', 'Hum', 'Pres']:
        if acumulado_ambientales[var]:
            df_final[var] = pd.concat(acumulado_ambientales[var], axis=1).mean(axis=1).round(2)

    for sensor in sensores_gas:
        if sensor in acumulado_gases and acumulado_gases[sensor]:
            df_final[sensor] = pd.concat(acumulado_gases[sensor], axis=1).mean(axis=1).round(2)

    return df_final

def obtener_fila_parametros_flat(df_datos, codigo_muestra):
    sensores_gas = [f"{L}{i}" for L in ['SA', 'SB', 'SC'] for i in range(1, 11)]
    fila = {'CÓDIGO': codigo_muestra}
    tiempo = df_datos['Tiempo'].values
    
    for sensor in sensores_gas:
        keys = [f"{sensor}_{suf}" for suf in ["Max", "T_Max", "Area", "Slope"]]
        if sensor not in df_datos.columns:
            for k in keys: fila[k] = None
            continue

        y_raw = df_datos[sensor].values
        mask = ~np.isnan(y_raw)
        if not np.any(mask):
            for k in keys: fila[k] = None
            continue
            
        y_clean = y_raw[mask]
        t_clean = tiempo[mask]
        
        max_val = np.max(y_clean)
        idx_max = np.argmax(y_clean)
        t_max = t_clean[idx_max]
        
        if hasattr(np, 'trapezoid'): area = np.trapezoid(y_clean, t_clean)
        else: area = np.trapz(y_clean, t_clean)
        
        slope = np.nan
        if idx_max + 5 <= len(y_clean):
            slope = np.polyfit(t_clean[idx_max:idx_max+5], y_clean[idx_max:idx_max+5], 1)[0]
            
        fila[keys[0]] = round(max_val, 4)
        fila[keys[1]] = round(t_max, 2)
        fila[keys[2]] = round(area, 4)
        fila[keys[3]] = round(slope, 6) if not pd.isna(slope) else None

    return fila

# 4. DIAGNÓSTICO DE CALIDAD CON LÓGICA DE PARES (PAIRWISE)

def analizar_calidad_completa(archivos_por_codigo):
    reporte_calidad = []
    print("\n   🕵️  Analizando Calidad (Filtrado Inteligente por Pares)...")

    for codigo, rutas in archivos_por_codigo.items():
        if len(rutas) < 2: continue
        
        dfs_reps = []
        rutas_validas = []
        
        for r in rutas:
            df_c = obtener_curvas_normalizadas_archivo(r)
            if df_c is not None:
                dfs_reps.append(df_c)
                rutas_validas.append(r)
                
        n_reps = len(dfs_reps)
        if n_reps < 2: continue
        
        sensores = [c for c in dfs_reps[0].columns if c != 'Tiempo']
        info_muestra = {'CÓDIGO': codigo, 'N_Rep_Iniciales': n_reps}
        
        # --- 1. CREAR MATRIZ DE SIMILITUD POR PARES ---
        sim_matrix = np.zeros((n_reps, n_reps))
        for i in range(n_reps):
            for j in range(i + 1, n_reps):
                corrs = []
                for s in sensores:
                    c1 = dfs_reps[i][s]
                    c2 = dfs_reps[j][s]
                    idx = c1.dropna().index.intersection(c2.dropna().index)
                    if len(idx) > 10:
                        try:
                            val = c1.loc[idx].corr(c2.loc[idx])
                            if not pd.isna(val): corrs.append(val)
                        except: pass
                avg_corr = np.mean(corrs) if corrs else 0
                sim_matrix[i, j] = avg_corr
                sim_matrix[j, i] = avg_corr
        
        # --- 2. DETECCIÓN DEL "CULPABLE" ---
        sugerencia = "Ninguna (Coherentes)"
        idx_descartar = -1
        
        if n_reps >= 3:
            mean_sims = []
            for i in range(n_reps):
                others = [sim_matrix[i, j] for j in range(n_reps) if i != j]
                mean_sims.append(np.mean(others))
            
            idx_peor = np.argmin(mean_sims)
            rem_idx = [i for i in range(n_reps) if i != idx_peor]
            rem_corrs = []
            for i in range(len(rem_idx)):
                for j in range(i + 1, len(rem_idx)):
                    rem_corrs.append(sim_matrix[rem_idx[i], rem_idx[j]])
                    
            avg_rem_corr = np.mean(rem_corrs) if rem_corrs else 0
            
            if avg_rem_corr >= 0.85 and mean_sims[idx_peor] < avg_rem_corr - 0.10:
                idx_descartar = idx_peor
                ruta_completa_outlier = rutas_validas[idx_peor]
                carpeta_padre = os.path.basename(os.path.dirname(ruta_completa_outlier))
                nombre_archivo = os.path.basename(ruta_completa_outlier)
                sugerencia = f"Eliminar: {carpeta_padre}/{nombre_archivo} (Salvando el resto)"
        
        elif n_reps == 2:
            if sim_matrix[0, 1] < 0.70:
                sugerencia = "Todas las reps son muy distintas"
                
        info_muestra['SUGERENCIA_DESCARTAR'] = sugerencia

        # --- 3. CÁLCULO DE MÉTRICAS GLOBALES EXCLUYENDO AL CULPABLE ---
        valid_reps_idx = [i for i in range(n_reps) if i != idx_descartar]
        
        scores_pearson = []
        scores_rsd = []
        
        if len(valid_reps_idx) >= 2:
            for sensor in sensores:
                # Pearson Limpio
                correlaciones_sensor = []
                for i in range(len(valid_reps_idx)):
                    for j in range(i + 1, len(valid_reps_idx)):
                        idx_i = valid_reps_idx[i]
                        idx_j = valid_reps_idx[j]
                        c1 = dfs_reps[idx_i][sensor]
                        c2 = dfs_reps[idx_j][sensor]
                        idx_validos = c1.dropna().index.intersection(c2.dropna().index)
                        if len(idx_validos) > 5:
                            try:
                                corr = c1.loc[idx_validos].corr(c2.loc[idx_validos])
                                if not pd.isna(corr): correlaciones_sensor.append(corr)
                            except: pass
                
                val_pearson = np.mean(correlaciones_sensor) if correlaciones_sensor else np.nan
                if not np.isnan(val_pearson): scores_pearson.append((sensor, val_pearson))

                # RSD Limpio
                maximos_reps = []
                for idx in valid_reps_idx:
                    df_r = dfs_reps[idx]
                    if sensor in df_r.columns:
                        max_val = df_r[sensor].max()
                        if not pd.isna(max_val): maximos_reps.append(max_val)
                
                if len(maximos_reps) >= 2:
                    media = np.mean(maximos_reps)
                    std = np.std(maximos_reps, ddof=1)
                    rsd = (std / media * 100) if media != 0 else 0
                    scores_rsd.append((sensor, abs(rsd)))

        # --- 4. DIAGNÓSTICO FINAL ---
        if scores_pearson and scores_rsd:
            scores_pearson.sort(key=lambda x: x[1])
            worst_sensor_pearson, val_min_pearson = scores_pearson[0]
            avg_pearson = np.mean([x[1] for x in scores_pearson])
            
            scores_rsd.sort(key=lambda x: x[1], reverse=True)
            worst_sensor_rsd, val_max_rsd = scores_rsd[0]
            avg_rsd = np.mean([x[1] for x in scores_rsd])

            info_muestra['Pearson_Global'] = round(avg_pearson, 4)
            info_muestra['Pearson_Min'] = round(val_min_pearson, 4)
            info_muestra['RSD_Global'] = round(avg_rsd, 2)
            
            estado = ""
            accion = ""
            
            if idx_descartar != -1:
                estado = "🟢 ÓPTIMO (Salvado por filtro)"
                accion = "Ver col. SUGERENCIA"
            elif sugerencia == "Todas las reps son muy distintas":
                estado = "🔴 CRÍTICO (IRRECUPERABLE)"
                accion = "Excluir muestra"
            elif avg_pearson < 0.70:
                estado = "🔴 CRÍTICO (FORMA)"
                accion = "Revisar todo"
            elif avg_rsd > 25.0:
                estado = "🟠 DIVERGENCIA INTENSIDAD"
                accion = "Posible pipeteo"
            elif val_min_pearson < 0.80:
                estado = "🟡 FALLO SENSOR LOCAL"
                accion = f"Ignorar {worst_sensor_pearson}"
            else:
                estado = "🟢 ÓPTIMO"
                accion = "Usar Todo"

            info_muestra['DIAGNÓSTICO'] = estado
            info_muestra['ACCION'] = accion
            reporte_calidad.append(info_muestra)
        else:
            info_muestra['DIAGNÓSTICO'] = "⚪ DATOS INSUFICIENTES"
            info_muestra['ACCION'] = "Revisar CSV"
            reporte_calidad.append(info_muestra)
            
    return pd.DataFrame(reporte_calidad)

# =============================================================================
# 5. MÓDULOS ANALÍTICOS Y GRÁFICOS (PCA E IMPACTOS)
# =============================================================================

def ejecutar_pca(df_params, nombre_proyecto, writer):
    print("\n   🧠 Ejecutando PCA (Análisis de Componentes Principales)...")
    
    df_pca = df_params.copy()
    if 'CÓDIGO' not in df_pca.columns: return
    
    codigos = df_pca['CÓDIGO']
    features = df_pca.drop(columns=['CÓDIGO'])
    features = features.fillna(0)
    
    scaler = StandardScaler()
    scaled_data = scaler.fit_transform(features)
    
    pca = PCA(n_components=2)
    principal_components = pca.fit_transform(scaled_data)
    
    varianza = pca.explained_variance_ratio_
    var_total = sum(varianza) * 100
    print(f"      -> Varianza explicada: {var_total:.2f}%")
    
    df_resultado_pca = pd.DataFrame(data=principal_components, columns=['PC1', 'PC2'])
    df_resultado_pca['CÓDIGO'] = codigos.values
    
    df_resultado_pca['GRUPO'] = df_resultado_pca['CÓDIGO'].apply(
        lambda x: re.match(r"([a-zA-Z]+)", str(x)).group(1) if re.match(r"([a-zA-Z]+)", str(x)) else "Desconocido"
    )
    
    grupos_detectados = df_resultado_pca['GRUPO'].unique()
    paleta_colores = {}
    for g in grupos_detectados:
        if g == 'VE': paleta_colores[g] = '#2ca02c'
        elif g == 'V': paleta_colores[g] = '#FFD700'
        else: paleta_colores[g] = '#d62728'
        
    df_resultado_pca.to_excel(writer, sheet_name='PCA_Coordenadas', index=False)
    
    plt.figure(figsize=(12, 9))
    sns.scatterplot(
        x='PC1', y='PC2', data=df_resultado_pca, s=150,                
        hue='GRUPO', palette=paleta_colores, alpha=0.8, edgecolor='black', linewidth=0.5
    )
    
    for i in range(df_resultado_pca.shape[0]):
        plt.text(
            df_resultado_pca.PC1[i]+0.2, df_resultado_pca.PC2[i], 
            df_resultado_pca.CÓDIGO[i], horizontalalignment='left', 
            size='small', color='#333333', weight='normal'
        )
        
    plt.title(f'PCA - Mapa de Olores ({nombre_proyecto})\nVarianza Explicada: {var_total:.1f}%')
    plt.xlabel(f'Componente Principal 1 ({varianza[0]*100:.1f}%)')
    plt.ylabel(f'Componente Principal 2 ({varianza[1]*100:.1f}%)')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend(title='Tipo de Muestra', fontsize='medium')
    plt.tight_layout()
    plt.savefig(f"PCA_Plot_{nombre_proyecto}.png", dpi=300)
    plt.close()
    
    print("      -> Gráfico PCA guardado.")

    # --- EXTRAER LOADINGS (PESOS) ---
    loadings = pd.DataFrame(
        pca.components_.T, 
        columns=['PC1', 'PC2'], 
        index=features.columns
    )
    loadings.to_excel(writer, sheet_name='PCA_Loadings')

    # --- TODOS LOS PARÁMETROS EN BARRAS HORIZONTALES ---
    for idx, pc in enumerate(loadings.columns):
        # Tomamos TODOS los parámetros con sus signos reales.
        all_pc = loadings[pc].sort_values(ascending=True)
        
        # Asignamos los colores respetando el orden de all_pc
        colores_barras = [asignar_color_parametro(var) for var in all_pc.index]
        
        # Ajustar la altura de la imagen
        alto_figura = max(10, len(all_pc) * 0.25)
        
        plt.figure(figsize=(12, alto_figura))
        
        # Graficamos en horizontal usando width en lugar de height para Pandas
        all_pc.plot(kind='barh', color=colores_barras, edgecolor='black', width=0.7)
        
        # Línea en el eje cero para ver claramente la polaridad
        plt.axvline(x=0, color='red', linestyle='-', linewidth=1.5, alpha=0.8)
        
        plt.title(f'Cargas (Loadings) Reales de TODOS los factores en {pc} ({varianza[idx]*100:.1f}% Varianza)', fontsize=14, fontweight='bold')
        plt.xlabel('Valor de la Carga (Loading)', fontsize=12, fontweight='bold')
        plt.ylabel('Parámetros', fontsize=12, fontweight='bold')
        
        plt.yticks(fontsize=8) 
        plt.grid(axis='x', linestyle='--', alpha=0.7)
        
        # Añadir leyenda de colores
        leyenda_verde = mpatches.Patch(color='#2ca02c', label='Área')
        leyenda_azul = mpatches.Patch(color='#1f77b4', label='T_Max')
        leyenda_naranja = mpatches.Patch(color='#ff7f0e', label='Max')
        leyenda_morado = mpatches.Patch(color='#9467bd', label='Slope')
        leyenda_gris = mpatches.Patch(color='#7f7f7f', label='Otros')
        
        plt.legend(handles=[leyenda_verde, leyenda_azul, leyenda_naranja, leyenda_morado, leyenda_gris], 
                   title='Tipos de Variable', loc='lower right')
        
        plt.tight_layout()
        plt.savefig(f"Importancia_Variables_{pc}_{nombre_proyecto}.png", dpi=300)
        plt.close()
        
        print(f"      -> Gráfico de cargas COMPLETO  para {pc} generado.")

def calcular_impacto_relativo(df_norm_filled, nombre_proyecto, writer):
    """
    Calcula el % de impacto (varianza) de cada parámetro sobre la distancia total
    para todos los pares posibles, genera la hoja Excel y el gráfico de barras horizontales.
    """
    print("\n   🎯 Calculando Impacto Relativo de Variables...")
    
    features = df_norm_filled.drop(columns=['CÓDIGO'])
    parametros = features.columns.tolist()
    n_muestras = len(df_norm_filled)
    
    matriz_impactos = []
    
    # 1. Calcular para todos los pares posibles
    for i in range(n_muestras):
        for j in range(i + 1, n_muestras):
            vec_a = features.iloc[i].values
            vec_b = features.iloc[j].values
            
            # Diferencia al cuadrado (Aportación pura)
            diff_sq = (vec_a - vec_b) ** 2
            dist_sq_total = np.sum(diff_sq)
            
            if dist_sq_total > 0:
                impacto_porcentual = (diff_sq / dist_sq_total) * 100.0
            else:
                impacto_porcentual = np.zeros_like(diff_sq)
                
            matriz_impactos.append(impacto_porcentual)
            
    if not matriz_impactos:
        print("      ⚠️ No se pudieron calcular impactos.")
        return
        
    # 2. Promedios y Desviaciones
    df_impactos = pd.DataFrame(matriz_impactos, columns=parametros)
    impacto_medio = df_impactos.mean()
    impacto_std = df_impactos.std(ddof=1)
    
    # Montar DataFrame ordenado de mayor a menor
    df_resumen_impacto = pd.DataFrame({
        'Parametro': parametros,
        'Impacto_Medio_Perc': impacto_medio.values,
        'Desviacion_Std_Perc': impacto_std.values
    }).sort_values(by='Impacto_Medio_Perc', ascending=False).reset_index(drop=True)
    
    # 3. Guardar en Excel
    df_resumen_impacto.to_excel(writer, sheet_name='Impacto_Variables', index=False)
    print("      -> Hoja 'Impacto_Variables' añadida al Excel.")
    
    # --- 4. GRÁFICO CON TODOS LOS PARÁMETROS ---
    # Ordenamos de forma ascendente para que el de mayor impacto quede arriba en el plot
    df_plot = df_resumen_impacto.sort_values(by='Impacto_Medio_Perc', ascending=True) 

    colores_barras = df_plot['Parametro'].apply(asignar_color_parametro)
    
    # Ajustar altura de la imagen dinámicamente según la cantidad de parámetros
    alto_figura = max(10, len(parametros) * 0.25)
    
    plt.figure(figsize=(12, alto_figura))
    plt.barh(df_plot['Parametro'], df_plot['Impacto_Medio_Perc'], 
             xerr=df_plot['Desviacion_Std_Perc'], capsize=2, 
             color=colores_barras, edgecolor='black', alpha=0.9, height=0.7)
    
    plt.xlabel('Impacto Relativo Medio (%)', fontsize=12, fontweight='bold')
    plt.title('Impacto Relativo de TODAS las Variables en la Diferenciación Global\n(Líneas negras representan Desviación Típica)', fontsize=14, fontweight='bold')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    
    # Reducir el tamaño de la fuente para que entren todos los nombres sin pisarse
    plt.yticks(fontsize=8)
    
    # Leyenda manual
    leyenda_verde = mpatches.Patch(color='#2ca02c', label='Área')
    leyenda_azul = mpatches.Patch(color='#1f77b4', label='T_Max')
    leyenda_naranja = mpatches.Patch(color='#ff7f0e', label='Max')
    leyenda_morado = mpatches.Patch(color='#9467bd', label='Slope')
    leyenda_gris = mpatches.Patch(color='#7f7f7f', label='Otros')

    plt.legend(handles=[leyenda_verde, leyenda_azul, leyenda_naranja, leyenda_morado, leyenda_gris], 
               loc='lower right', title='Tipos de Variable')

    plt.tight_layout()
    plt.savefig(f"Impacto_Relativo_Var_{nombre_proyecto}.png", dpi=300)
    plt.close()
    
    print("      -> Gráfico de Barras de Impacto (TODOS los parámetros) generado con éxito.")
# 6. MAIN

def main():
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    os.chdir(ruta_script)
    print("="*60)
    print("🔬 PROCESADOR FINAL (CON FILTRO INTELIGENTE DE ERRORES)")
    print("="*60)
    nombre_carpeta_general = input("📂 Introduce la carpeta general: ").strip()
    ruta_general = os.path.join(ruta_script, nombre_carpeta_general)
    if not os.path.exists(ruta_general):
        print("❌ Carpeta no encontrada.")
        return
    archivos_por_codigo = {} 
    lista_resumen = [] 
    for raiz, _, archivos in os.walk(ruta_general):
        for archivo in archivos:
            if archivo.endswith(".csv"):
                ruta_completa = os.path.join(raiz, archivo)
                meta = extraer_metadatos_csv(ruta_completa)
                meta["CARPETA"] = os.path.basename(raiz)
                lista_resumen.append(meta)
                if meta["CÓDIGO"]:
                    if meta["CÓDIGO"] not in archivos_por_codigo: archivos_por_codigo[meta["CÓDIGO"]] = []
                    archivos_por_codigo[meta["CÓDIGO"]].append(ruta_completa)
    if not lista_resumen: return
    # FASE 1: CALIDAD
    df_calidad = analizar_calidad_completa(archivos_por_codigo)
    archivos_blacklist = []
    sensores_blacklist_map = {}
    if not df_calidad.empty:
        conteo = df_calidad['DIAGNÓSTICO'].value_counts()
        print("\n📊 RESUMEN DE CALIDAD DE MUESTRAS:")
        for estado, count in conteo.items():
            print(f"   {estado}: {count} muestras")
        for idx, row in df_calidad.iterrows():
            sug = str(row.get('SUGERENCIA_DESCARTAR', ''))
            acc = str(row.get('ACCION', ''))
            cod = row['CÓDIGO']
            if "Eliminar:" in sug:
                try:
                    parte_ruta = sug.split("Eliminar:")[1].split("(")[0].strip()
                    if parte_ruta:
                        archivos_blacklist.append(parte_ruta)
                        print(f"      🚫 Se excluirá repetición: {parte_ruta}")
                except: pass
            if acc.startswith("Ignorar "):
                sensor_bad = acc.split("Ignorar ")[1].strip()
                if cod not in sensores_blacklist_map: sensores_blacklist_map[cod] = []
                sensores_blacklist_map[cod].append(sensor_bad)
                print(f"      🔇 Se silenciará sensor {sensor_bad} para muestra {cod}")
        cols_prin = ['CÓDIGO', 'N_Rep_Iniciales', 'DIAGNÓSTICO', 'ACCION', 'SUGERENCIA_DESCARTAR',
                     'Pearson_Global', 'Pearson_Min', 'RSD_Global']
        cols_rest = sorted([c for c in df_calidad.columns if c not in cols_prin])
        df_calidad = df_calidad[cols_prin + cols_rest]
        df_calidad['temp'] = df_calidad['CÓDIGO'].apply(clave_orden_natural)
        df_calidad = df_calidad.sort_values(by='temp').drop(columns=['temp'])
    # FASE 2: PROCESAMIENTO
    nombre_excel = f"Analisis_Completo_{nombre_carpeta_general}.xlsx"
    print(f"\n💾 Generando Excel: {nombre_excel}...")
    lista_filas_planas = []
    matriz_global_dist = None 
    try:
        with pd.ExcelWriter(nombre_excel, engine='openpyxl') as writer:
            df_resumen = pd.DataFrame(lista_resumen)
            if "CÓDIGO" in df_resumen.columns:
                df_resumen['temp'] = df_resumen['CÓDIGO'].apply(clave_orden_natural)
                df_resumen = df_resumen.sort_values(by='temp').drop(columns=['temp'])
            df_resumen.to_excel(writer, sheet_name='Resumen_Archivos', index=False)
            if not df_calidad.empty:
                df_calidad.to_excel(writer, sheet_name='Diagnostico_Calidad', index=False)
            codigos_ordenados = sorted(archivos_por_codigo.keys(), key=lambda x: clave_orden_natural(x))
            hojas_curvas = {}
            print("   ⚙️  Calculando parámetros promedio (aplicando filtros)...")
            for codigo in codigos_ordenados:
                lista_archivos_raw = archivos_por_codigo[codigo]
                lista_archivos_limpia = []
                for ruta in lista_archivos_raw:
                    identificador = f"{os.path.basename(os.path.dirname(ruta))}/{os.path.basename(ruta)}"
                    if identificador not in archivos_blacklist:
                        lista_archivos_limpia.append(ruta)
                df_curva_promedio = procesar_grupo_ensayos(lista_archivos_limpia) 
                if not df_curva_promedio.empty:
                    hojas_curvas[codigo] = df_curva_promedio
                    lista_filas_planas.append(obtener_fila_parametros_flat(df_curva_promedio, codigo))
            df_dist = pd.DataFrame()
            df_params = pd.DataFrame() 
            if lista_filas_planas:
                df_params = pd.DataFrame(lista_filas_planas)
                lista_sensores = [f"{L}{i}" for L in ['SA','SB','SC'] for i in range(1,11)]
                cols_ordenadas = ['CÓDIGO']
                for s in lista_sensores: cols_ordenadas.extend([f"{s}_Max", f"{s}_T_Max", f"{s}_Area", f"{s}_Slope"])
                df_params = df_params[[c for c in cols_ordenadas if c in df_params.columns]]
                for cod_bad, sensores_bad in sensores_blacklist_map.items():
                    indices = df_params.index[df_params['CÓDIGO'] == cod_bad].tolist()
                    for idx in indices:
                        for s_bad in sensores_bad:
                            cols_to_mask = [c for c in df_params.columns if c.startswith(f"{s_bad}_")]
                            if cols_to_mask:
                                df_params.loc[idx, cols_to_mask] = 0.0
                if not df_params.empty and len(df_params) > 2:
                    ejecutar_pca(df_params, nombre_carpeta_general, writer)
                print("   ⚖️  Calculando Normalización...")
                df_norm = df_params.copy()
                for col in df_norm.columns.drop('CÓDIGO'):
                    if df_norm[col].dtype in [float, int]:
                        min_v, max_v = df_norm[col].min(), df_norm[col].max()
                        if max_v > min_v: df_norm[col] = (df_norm[col] - min_v) / (max_v - min_v)
                        else: df_norm[col] = 0.0
                        df_norm[col] = df_norm[col].round(4)
                df_norm_filled = df_norm.fillna(0)
                
                # --- LLAMADA A LA NUEVA FUNCIÓN DE IMPACTO RELATIVO ---
                calcular_impacto_relativo(df_norm_filled, nombre_carpeta_general, writer)
                print("   📐 Calculando Distancias Normalizadas...")
                filas_distancia = []
                codigos = df_norm_filled['CÓDIGO'].values
                n = len(codigos)
                matriz_global_vals = np.zeros((n, n))
                for i in range(n):
                    for j in range(i + 1, n):
                        res = {'PAR': f"{codigos[i]} vs {codigos[j]}"}
                        vec_a = df_norm_filled.iloc[i]
                        vec_b = df_norm_filled.iloc[j]
                        dist_acumulada_par = []
                        for s in lista_sensores:
                            cols_s = [c for c in df_norm_filled.columns if c.startswith(f"{s}_")]
                            if cols_s:
                                va = vec_a[cols_s].values.astype(float)
                                vb = vec_b[cols_s].values.astype(float)
                                d = np.linalg.norm(va - vb) / np.sqrt(len(cols_s))
                                res[s] = round(d, 4)
                                dist_acumulada_par.append(d)
                            else: res[s] = None
                        if dist_acumulada_par:
                            dist_media = np.mean(dist_acumulada_par)
                            matriz_global_vals[i, j] = dist_media
                            matriz_global_vals[j, i] = dist_media
                        filas_distancia.append(res)
                if filas_distancia:
                    df_dist = pd.DataFrame(filas_distancia)
                    df_dist.to_excel(writer, sheet_name='Distancias_Pares', index=False)
                    matriz_global_dist = pd.DataFrame(matriz_global_vals, index=codigos, columns=codigos)

            for c, df_d in hojas_curvas.items():
                df_d.to_excel(writer, sheet_name=re.sub(r'[\\/*?:\[\]]', '_', str(c)[:31]), index=False)
            if 'Diagnostico_Calidad' in writer.sheets:
                ws_calidad = writer.sheets['Diagnostico_Calidad']
                aplicar_colores_rsd(ws_calidad)      
                aplicar_colores_pearson(ws_calidad)  
            for sheet_name in writer.sheets:
                auto_ajustar_columnas(writer.sheets[sheet_name])
                
        # --- AQUÍ CORREGÍ TU ERROR DE INDENTACIÓN AL DIBUJAR LA MATRIZ DE CALOR ---
        if matriz_global_dist is not None:
            plt.figure(figsize=(12, 10))
            sns.heatmap(matriz_global_dist, cmap='viridis', linewidths=0.05, linecolor='white', cbar_kws={'label': 'Distancia Promedio Global'})
            plt.title("Matriz de Similitud Global")
            plt.xticks(rotation=90, fontsize=8)
            plt.yticks(rotation=0, fontsize=8)
            plt.tight_layout()
            plt.savefig(f"Matriz_Global_{nombre_carpeta_general}.png", dpi=300)
            plt.close()

        print(f"\n✅ ¡TODO LISTO! Excel coloreado y gráficos generados.")

    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
    input("\nPresiona ENTER para salir...")

if __name__ == "__main__":
    main()