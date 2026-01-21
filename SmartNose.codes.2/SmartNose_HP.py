import serial
import matplotlib.pyplot as plt
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import time
import platform
import os
import numpy as np


# ---------------- CONFIGURACIÓN ----------------
PUERTO = "COM14"
BAUDIOS = 115200
TIMEOUT = 0.1
# ------------------------------------------------

# Forzar carpeta del script como working dir
RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
os.chdir(RUTA_BASE)
print(f"📂 Archivos se guardarán en: {RUTA_BASE}")

def beep(n=1):
    try:
        if platform.system() == "Windows":
            import winsound
            for _ in range(n):
                winsound.Beep(1000, 500)
                time.sleep(0.1)
        else:
            for _ in range(n):
                print("\a")
                time.sleep(0.1)
    except Exception:
        pass

def teclado_stop(ser):
    try:
        import msvcrt
        if msvcrt.kbhit():
            key = msvcrt.getch().decode().lower()
            if key == 's':
                ser.write(b"STOP\n")
                print("⛔ STOP enviado al ESP32 (teclado).")
                return True
    except ImportError:
        pass
    return False

def esperar_ready(ser):
    print("🔄 Esperando conexión con ESP32...")
    while True:
        linea = ser.readline().decode(errors="ignore").strip()
        if linea:
            print("ESP32:", linea)
            if linea == "READY":
                print("✅ ESP32 listo.")
                break

def esperar_conf(ser):
    while True:
        linea = ser.readline().decode(errors="ignore").strip()
        if linea == "CONF":
            print("✅ Arduino listo para recibir configuración.")
            break

def normalizar_camara(s):
    s = s.strip().lower()
    if s in ("a", "abierta"): return "abierta"
    if s in ("c", "cerrada"): return "cerrada"
    return None

def estadisticas(data):
    if len(data) == 0:
        return {"media": np.nan, "std": np.nan, "var": np.nan, "min": np.nan, "max": np.nan}
    return {
        "media": float(np.mean(data)),
        "std": float(np.std(data, ddof=0)),
        "var": float(np.var(data, ddof=0)),
        "min": float(np.min(data)),
        "max": float(np.max(data))
    }
import matplotlib.pyplot as plt

def actualizar_grafico_progreso(tiempo_transcurrido, tiempo_total, fase_nombre):
    """
    Muestra un gráfico de pastel circular que representa el progreso del tiempo en la fase.
    - tiempo_transcurrido: segundos pasados (float).
    - tiempo_total: duración total de la fase en segundos (float).
    - fase_nombre: nombre de la fase (str).
    """
    if tiempo_total <= 0:
        return  # Evitar división por cero
    
    # Calcular porcentajes
    porcentaje_transcurrido = min((tiempo_transcurrido / tiempo_total) * 100, 100)
    porcentaje_restante = 100 - porcentaje_transcurrido
    
    # Datos para el pie chart
    sizes = [porcentaje_transcurrido, porcentaje_restante]
    labels = [f'Transcurrido: {tiempo_transcurrido:.1f}s', f'Restante: {tiempo_total - tiempo_transcurrido:.1f}s']
    colors = ['blue', 'lightgray']  # Azul para transcurrido, gris para restante
    
    # Crear o actualizar el gráfico
    plt.figure(1)  # Usar figura 1 para reutilizar la ventana
    plt.clf()  # Limpiar la figura anterior
    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    plt.title(f'Progreso de la Fase: {fase_nombre}\nTiempo Total: {tiempo_total:.1f}s')
    plt.axis('equal')  # Hacerlo circular
    plt.pause(0.1)  # Pausa breve para actualizar la gráfica (sin bloquear)


def ejecutar_ensayo(ser):
    nombre_ensayo = input("\nNombre del ensayo: ")
    fecha_inicio = datetime.now()
    fecha_str = fecha_inicio.strftime("%Y-%m-%d_%H-%M-%S")

    # Pedir número de fases y configuraciones
    while True:
        try:
            num_fases = int(input("Número de fases: "))
            if num_fases < 1:
                print("Debe ser al menos 1.")
                continue
            break
        except ValueError:
            print("Por favor ingresa un número entero válido.")
    fases = []
    for i in range(num_fases):
        fase_id = f"F.{i+1}"
        print(f"\nConfiguración fase {i+1} (ID: {fase_id}):")
        tipo_muestra = input("Muestra: ").strip()
        while not tipo_muestra:
            print("La muestra no puede estar vacía.")
            tipo_muestra = input("Muestra: ").strip()
        while True:
            cam = input("Cámara (a=abierta, c=cerrada): ").strip()
            camara = normalizar_camara(cam)
            if camara is None:
                print("Por favor ingresa 'a' o 'c' (o 'abierta'/'cerrada').")
            else:
                break
        while True:
            try:
                dur = float(input("Duración (minutos): "))
                if dur <= 0:
                    print("Debe ser positivo.")
                    continue
                break
            except ValueError:
                print("Ingresa un número válido.")
        while True:
            try:
                vel = int(input("Velocidad ventilador (0-100): "))
                vel = max(0, min(100, vel))
                break
            except ValueError:
                print("Ingresa un entero 0-100.")
        fases.append({
            "nombre": fase_id,
            "tipo_muestra": tipo_muestra,
            "camara_estado": camara,
            "duracion": dur,
            "ventilador": vel
        })

    # Variables 
    variables = (
        ["Temp_SA", "Temp_SB", "Temp_SC",
         "Hum_SA", "Hum_SB", "Hum_SC",
         "Pres_SA", "Pres_SB", "Pres_SC"]
        + [f"SA{i+1}" for i in range(10)]
        + [f"SB{i+1}" for i in range(10)]
        + [f"SC{i+1}" for i in range(10)]
    )

    # Estructuras para datos
    datos_t = []
    datos_pwm = []
    datos_fase = []
    datos_vars = [[] for _ in range(len(variables))]  # 12 listas
    tiempo_acumulado = 0

    # Loop por fases
    for idx_fase, fase in enumerate(fases):
        print(f"\nConfigurando fase {idx_fase+1}: {fase['nombre']}")
        print(f"Muestra: {fase['tipo_muestra']}")
        print(f"Cámara: {fase['camara_estado']}")
        print(f"Duración (minutos): {fase['duracion']}")
        print(f"Velocidad ventilador (%): {fase['ventilador']}")

        ser.write(f"{nombre_ensayo}\n".encode())
        esperar_conf(ser)

        ser.write(f"{fase['tipo_muestra']}\n".encode()); time.sleep(0.05)
        ser.write(f"{fase['camara_estado']}\n".encode()); time.sleep(0.05)
        ser.write(f"{int(fase['duracion']*60)}\n".encode()); time.sleep(0.05)
        ser.write(f"{fase['ventilador']}\n".encode()); time.sleep(0.05)

        if idx_fase == 0:
            input(f"Presiona ENTER para iniciar la fase '{fase['nombre']}'... ")

        else:
            input(f"Presiona ENTER para iniciar la fase '{fase['nombre']}'... ")
        ser.write(b"ENTER\n")

        print(f"⏳ Medición en curso en fase '{fase['nombre']}'... (pulsa S para STOP)")

        fase_terminada = False
        try:
            while True:
                if teclado_stop(ser):
                    ser.write(b"STOP\n")
                    print("⛔ STOP enviado al ESP32 (teclado).")
                    fase_terminada = True
                    break

                linea = ser.readline().decode(errors="ignore").strip()
                if not linea:
                    plt.pause(0.01)
                    continue

                if linea.startswith("DATA"):
                    campos = linea.split("|")
                    EXPECTED_FIELDS = 42
                    if len(campos) != EXPECTED_FIELDS:
                        print(f"⚠️ Línea DATA ignorada: número de campos incorrecto ({len(campos)} != {EXPECTED_FIELDS}).")
                        print(f"Línea recibida: {linea}")
                        continue
                    try:
                        t_relativo =int(campos[1])
                        t = t_relativo + tiempo_acumulado
                        valores = [float(val) for val in campos[2:-1]]  # desde campos[2] hasta el penúltimo
                        pwm = int(campos[-1])  # último campo
                    except ValueError as e:
                        print(f"⚠️ Error al convertir a número en línea DATA: {e}. Línea: {linea}")
                        continue

                    # Guardar variables
                    for i_v, v in enumerate(valores):
                        if i_v < len(datos_vars):
                            datos_vars[i_v].append(v)

                    datos_t.append(t)
                    datos_pwm.append(pwm)
                    datos_fase.append(fase["nombre"])


                    # Escribir fila raw en CSV
                    fila = [fase["nombre"], t, pwm] + valores 
                    

                    # Calcular tiempo transcurrido para la fase actual
                    tiempo_total_fase = fase['duracion'] * 60  # Duración en segundos
                    tiempo_transcurrido_fase = t_relativo  # t_relativo es el tiempo relativo desde el inicio de la fase

                    # Actualizar gráfico de progreso
                    actualizar_grafico_progreso(tiempo_transcurrido_fase, tiempo_total_fase, fase["nombre"])

                elif linea == "FIN":
                    print(f"✅ Fase '{fase['nombre']}' finalizada")
                    beep()
                    fase_terminada = True
                    break

                elif linea == "STOPPED":
                    print(f"🟥 Fase '{fase['nombre']}' detenida manualmente")
                    fase_terminada = True
                    break

            if not fase_terminada:
                print(f"⚠️ Fase '{fase['nombre']}' terminó inesperadamente")

        except Exception as e:
            print(f"Error durante la fase '{fase['nombre']}': {e}")
            fase_terminada = True

        tiempo_acumulado = datos_t[-1] if datos_t else tiempo_acumulado

        if idx_fase < len(fases) - 1:
            input("Presiona ENTER para continuar a la siguiente fase...")


    

    # Crear CSV final con columnas extra de suavizado
    nombre_csv_final = f"{nombre_ensayo}_{fecha_str}.csv"
    with open(nombre_csv_final, "w", newline="", encoding="utf-8") as ffinal:
        writer = csv.writer(ffinal)
        header_final = ["Fase", "Tiempo(s)", "PWM"] + variables 
        writer.writerow(header_final)
        n = len(datos_t)
        for i in range(n):
            row = [datos_fase[i], datos_t[i], datos_pwm[i]]
            for j in range(len(variables)):
                val = datos_vars[j][i]
                if np.isnan(val):
                    row.append("")  # Blanco para NaN
                else:
                    row.append(val)
            writer.writerow(row)
    print(f"📁 CSV final con suavizado guardado en: {os.path.abspath(nombre_csv_final)}")

    # --- Cálculos estadísticos y ajustes (para hoja Parámetros) ---
    # Convertir datos_vars a array (12 x N)
    datos_np = np.array(datos_vars) if len(datos_vars) > 0 else np.zeros((len(variables), 0))
    xdata = np.array(datos_t)

    # medias por timestamp (promedio sensores)
    temp_media_ts = np.mean(datos_np[0:3, :], axis=0) if datos_np.shape[1] > 0 else np.array([])
    hum_media_ts = np.mean(datos_np[3:6, :], axis=0) if datos_np.shape[1] > 0 else np.array([])  # Promedio de Hum_A, Hum_B, Hum_C por timestamp
    pres_media_ts = np.mean(datos_np[6:9, :], axis=0) if datos_np.shape[1] > 0 else np.array([])


    # Estadísticas (usar series suavizadas si existen, si no las crudas)
    stats_hum = estadisticas(hum_media_ts)
    stats_temp = estadisticas(temp_media_ts)
    stats_pres = estadisticas(pres_media_ts)

    # --- Crear Excel con hojas: Parámetros, Datos, Gráficas ---
    nombre_excel = f"{nombre_ensayo}_{fecha_str}.xlsx"
    wb = Workbook()

    # Hoja "Parámetros" 
    ws_param = wb.active
    ws_param.title = "Parámetros"
    ws_param.append(["Nombre ensayo", nombre_ensayo])
    ws_param.append(["Fecha inicio", fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")])
    ws_param.append([])

    # Fases: encabezado y lista
    ws_param.append(["Fase", "Muestra", "Cámara", "Duración (min)", "Ventilador (%)"])
    for f in fases:
        ws_param.append([f["nombre"], f["tipo_muestra"], f["camara_estado"], f["duracion"], f["ventilador"]])
    ws_param.append([])

    ws_param.append(["Variable", "Media", "Std", "Var", "Min", "Max"])
   
    # Humedad
    ws_param.append([
        "Humedad",
        stats_hum["media"], stats_hum["std"], stats_hum["var"], stats_hum["min"], stats_hum["max"]
    ])
    # Temperatura 
    ws_param.append([
        "Temperatura",
        stats_temp["media"], stats_temp["std"], stats_temp["var"], stats_temp["min"], stats_temp["max"]
    ])
    # Presión 
    ws_param.append([
        "Presion",
        stats_pres["media"], stats_pres["std"], stats_pres["var"], stats_pres["min"], stats_pres["max"]
    ])

    # Hoja "Datos"
    ws_datos = wb.create_sheet(title="Datos")
    header_excel = ["Fase", "Tiempo(s)", "PWM"] + variables 
    ws_datos.append(header_excel)
    unidades = ["", "s", ""] + ["°C","°C","°C","%","%","%","hPa","hPa","hPa"] + ["kΩ"]*30
    ws_datos.append(unidades)

    N = len(datos_t)
    for i in range(N):
        row = [datos_fase[i], datos_t[i], datos_pwm[i]]
        row += [datos_vars[j][i] for j in range(len(variables))]
        ws_datos.append(row)

    # Hoja "Gráficas"
    fig_final, axs_final = plt.subplots(3, 1, figsize=(12, 9), sharex=True)  # Cambiado a 3 filas
    fig_final.suptitle(f"Ensayo: {nombre_ensayo}")
    
    # --- Humedad ---
    axs_final[0].plot(datos_t, datos_np[3], 'b-', alpha=0.8, label='Hum_A raw')  # Hum_A
    axs_final[0].plot(datos_t, datos_np[4], 'r-', alpha=0.8, label='Hum_B raw')  # Hum_B
    axs_final[0].plot(datos_t, datos_np[5], 'g-', alpha=0.8, label='Hum_C raw')  # Hum_C
    axs_final[0].plot(datos_t, hum_media_ts, 'y-', linewidth=2, label='Hum Media')  # Media
    axs_final[0].set_ylabel("Humedad (%)")
    axs_final[0].legend()
    axs_final[0].grid(True)

    # --- Temperatura ---
    axs_final[1].plot(datos_t, datos_np[0], 'b-', alpha=0.8, label='Temp_A raw')  # Temp_A
    axs_final[1].plot(datos_t, datos_np[1], 'r-', alpha=0.8, label='Temp_B raw')  # Temp_B
    axs_final[1].plot(datos_t, datos_np[2], 'g-', alpha=0.8, label='Temp_C raw')  # Temp_C
    axs_final[1].plot(datos_t, temp_media_ts, 'y-', linewidth=2, label='Temp Media')  # Media
    axs_final[1].set_ylabel("Temperatura (°C)")
    axs_final[1].legend()
    axs_final[1].grid(True)

    # --- Presión ---
    axs_final[2].plot(datos_t, datos_np[6], 'b-', alpha=0.8, label='Pres_A raw')  # Pres_A
    axs_final[2].plot(datos_t, datos_np[7], 'r-', alpha=0.8, label='Pres_B raw')  # Pres_B
    axs_final[2].plot(datos_t, datos_np[8], 'g-', alpha=0.8, label='Pres_C raw')  # Pres_C
    axs_final[2].plot(datos_t, pres_media_ts, 'y-', linewidth=2, label='Pres Media')  # Media
    axs_final[2].set_ylabel("Presión (hPa)")
    axs_final[2].set_xlabel("Tiempo (s)")
    axs_final[2].legend()
    axs_final[2].grid(True)

    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    img_path = f"grafica_{fecha_str}.png"
    fig_final.savefig(img_path, bbox_inches="tight")
    plt.close(fig_final)  # cerrar figura para liberar memoria

    # Insertar en Excel
    ws_graf = wb.create_sheet(title="Gráficas")
    try:
        img = ExcelImage(img_path)
        img.anchor = 'A1'
        ws_graf.add_image(img)
    except Exception as e:
        print(f"⚠️ No se pudo insertar imagen en Excel: {e}")

    # Ordenar hojas
        wb._sheets = [ws_param, ws_datos, ws_graf]

    # Guardar workbook 
    try:
        wb.save(nombre_excel)
        print(f"📊 Excel guardado correctamente: {os.path.abspath(nombre_excel)}")
    except Exception as e:
        print(f"❌ Error al guardar Excel: {e}")
    # borrar png temporal si existe
    try:
        if os.path.exists(img_path):
            os.remove(img_path)
    except Exception:
        pass

    beep()
    plt.ioff()
    print("✅ Ensayo finalizado correctamente.")
    print("\n📂 Contenido del directorio actual:")
    print(os.listdir('.'))

def main():
    ser = None
    while True:
        try:
            if ser is None or not ser.is_open:
                ser = serial.Serial(PUERTO, BAUDIOS, timeout=TIMEOUT)
                time.sleep(2)
                ser.reset_input_buffer()
                esperar_ready(ser)
            ejecutar_ensayo(ser)
            print("\nEnsayo finalizado. Puedes iniciar otro ensayo o presionar Ctrl+C para salir.\n")
        except serial.SerialException as e:
            print(f"Error de conexión: {e}")
            print("Reconectando en 5 segundos...")
            if ser:
                try: ser.close()
                except: pass
            ser = None
            time.sleep(5)
        except KeyboardInterrupt:
            print("Programa terminado por usuario.")
            if ser:
                ser.close()
            break
        except Exception as e:
            print(f"Error inesperado: {e}")
            if ser:
                try: ser.close()
                except: pass
            ser = None
            time.sleep(5)

if __name__ == "__main__":
    main()
