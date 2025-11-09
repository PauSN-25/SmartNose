import serial
import matplotlib.pyplot as plt
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import time
import platform
import os
import numpy as np
from scipy.optimize import curve_fit
import warnings
from scipy.optimize import OptimizeWarning
warnings.filterwarnings("ignore", category=OptimizeWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)

# ---------------- CONFIGURACIÓN ----------------
PUERTO = "COM9"
BAUDIOS = 115200
TIMEOUT = 0.1
VENTANA_SUAVIZADO = 5
# ------------------------------------------------

# Forzar carpeta del script como working dir
RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
os.chdir(RUTA_BASE)
print(f"📂 Archivos se guardarán en: {RUTA_BASE}")

def beep(n=1):
    try:
        if platform.system() == "Windows":
            import winsound
            for _ in range(n):
                winsound.Beep(1000, 500)
                time.sleep(0.1)
        else:
            for _ in range(n):
                print("\a")
                time.sleep(0.1)
    except Exception:
        pass

def teclado_stop(ser):
    try:
        import msvcrt
        if msvcrt.kbhit():
            key = msvcrt.getch().decode().lower()
            if key == 's':
                ser.write(b"STOP\n")
                print("⛔ STOP enviado al ESP32 (teclado).")
                return True
    except ImportError:
        pass
    return False

def esperar_ready(ser):
    print("🔄 Esperando conexión con ESP32...")
    while True:
        linea = ser.readline().decode(errors="ignore").strip()
        if linea:
            print("ESP32:", linea)
            if linea == "READY":
                print("✅ ESP32 listo.")
                break

def esperar_conf(ser):
    while True:
        linea = ser.readline().decode(errors="ignore").strip()
        if linea == "CONF":
            print("✅ Arduino listo para recibir configuración.")
            break

def normalizar_camara(s):
    s = s.strip().lower()
    if s in ("a", "abierta"): return "abierta"
    if s in ("c", "cerrada"): return "cerrada"
    return None

def media_movil(arr, ventana=5):
    """Media móvil (modo 'same') para devolver arreglo de la misma longitud."""
    if len(arr) == 0:
        return np.array([])
    kernel = np.ones(ventana) / ventana
    return np.convolve(arr, kernel, mode='same')

def funcion_ajuste(x, y0, A, k):
    return y0 + A * (1 - np.exp(-k * x))

def calcular_r2(y_real, y_pred):
    ss_res = np.sum((y_real - y_pred) ** 2)
    ss_tot = np.sum((y_real - np.mean(y_real)) ** 2)
    return 1 - ss_res / ss_tot if ss_tot != 0 else np.nan

def estadisticas(data):
    if len(data) == 0:
        return {"media": np.nan, "mediana": np.nan, "std": np.nan, "var": np.nan, "min": np.nan, "max": np.nan}
    return {
        "media": float(np.mean(data)),
        "mediana": float(np.median(data)),
        "std": float(np.std(data, ddof=0)),
        "var": float(np.var(data, ddof=0)),
        "min": float(np.min(data)),
        "max": float(np.max(data))
    }

def ejecutar_ensayo(ser):
    nombre_ensayo = input("\nNombre del ensayo: ")
    fecha_inicio = datetime.now()
    fecha_str = fecha_inicio.strftime("%Y-%m-%d_%H-%M-%S")

    # Pedir número de fases y configuraciones
    while True:
        try:
            num_fases = int(input("Número de fases: "))
            if num_fases < 1:
                print("Debe ser al menos 1.")
                continue
            break
        except ValueError:
            print("Por favor ingresa un número entero válido.")
    fases = []
    for i in range(num_fases):
        fase_id = f"F.{i+1}"
        print(f"\nConfiguración fase {i+1} (ID: {fase_id}):")
        tipo_muestra = input("Muestra: ").strip()
        while not tipo_muestra:
            print("La muestra no puede estar vacía.")
            tipo_muestra = input("Muestra: ").strip()
        while True:
            cam = input("Cámara (a=abierta, c=cerrada): ").strip()
            camara = normalizar_camara(cam)
            if camara is None:
                print("Por favor ingresa 'a' o 'c' (o 'abierta'/'cerrada').")
            else:
                break
        while True:
            try:
                dur = float(input("Duración (minutos): "))
                if dur <= 0:
                    print("Debe ser positivo.")
                    continue
                break
            except ValueError:
                print("Ingresa un número válido.")
        while True:
            try:
                vel = int(input("Velocidad ventilador (0-100): "))
                vel = max(0, min(100, vel))
                break
            except ValueError:
                print("Ingresa un entero 0-100.")
        fases.append({
            "nombre": fase_id,
            "tipo_muestra": tipo_muestra,
            "camara_estado": camara,
            "duracion": dur,
            "ventilador": vel
        })

    # Variables 
    variables = [
        "Temp_A","Temp_B","Temp_C",
        "Hum_A","Hum_B","Hum_C",
        "Pres_A","Pres_B","Pres_C",
        "Gas_A","Gas_B","Gas_C"
    ]

    # Estructuras para datos
    datos_t = []
    datos_pwm = []
    datos_fase = []
    datos_vars = [[] for _ in range(len(variables))]  # 12 listas
    hum_medias = []
    gas_medias = []

    # Loop por fases
    for idx_fase, fase in enumerate(fases):
        print(f"\nConfigurando fase {idx_fase+1}: {fase['nombre']}")
        print(f"Muestra: {fase['tipo_muestra']}")
        print(f"Cámara: {fase['camara_estado']}")
        print(f"Duración (minutos): {fase['duracion']}")
        print(f"Velocidad ventilador (%): {fase['ventilador']}")

        ser.write(f"{nombre_ensayo}\n".encode())
        esperar_conf(ser)

        ser.write(f"{fase['tipo_muestra']}\n".encode()); time.sleep(0.05)
        ser.write(f"{fase['camara_estado']}\n".encode()); time.sleep(0.05)
        ser.write(f"{int(fase['duracion']*60)}\n".encode()); time.sleep(0.05)
        ser.write(f"{fase['ventilador']}\n".encode()); time.sleep(0.05)

        if idx_fase == 0:
            input(f"Presiona ENTER para iniciar la fase '{fase['nombre']}'... ")

            # Gráficas en tiempo real
            plt.ion()
            fig, axs = plt.subplots(4, 1, figsize=(10, 12), sharex=True)
            fig.suptitle(f"Ensayo: {nombre_ensayo}")
            lineas = []
            orden_graficas = [9,10,11, 3,4,5, 0,1,2, 6,7,8]
            titulos = ["Gas (kΩ)", "Humedad (%)", "Temperatura (°C)", "Presión (hPa)"]
            for i, ax in enumerate(axs):
                ax.set_ylabel(titulos[i])
                for j in range(3):
                    idx_var = orden_graficas[i*3 + j]
                    linea, = ax.plot([], [], label=variables[idx_var])
                    lineas.append(linea)
                ax.legend(loc="upper right")
            axs[-1].set_xlabel("Tiempo (s)")
            tiempo_acumulado = 0
            # ------------------------------------------------

        else:
            input(f"Presiona ENTER para iniciar la fase '{fase['nombre']}'... ")
        ser.write(b"ENTER\n")

        print(f"⏳ Medición en curso en fase '{fase['nombre']}'... (pulsa S para STOP)")

        fase_terminada = False
        try:
            while True:
                if teclado_stop(ser):
                    ser.write(b"STOP\n")
                    print("⛔ STOP enviado al ESP32 (teclado).")
                    fase_terminada = True
                    break

                linea = ser.readline().decode(errors="ignore").strip()
                if not linea:
                    plt.pause(0.01)
                    continue

                if linea.startswith("DATA"):
                    campos = linea.split("|")
                    if len(campos) != 15:
                        continue
                    t = int(campos[1]) + tiempo_acumulado
                    valores = list(map(float, campos[2:14]))  # 12 valores
                    pwm = int(campos[14])

                    # Guardar por-variable
                    for i_v, v in enumerate(valores):
                        datos_vars[i_v].append(v)

                    datos_t.append(t)
                    datos_pwm.append(pwm)
                    datos_fase.append(fase["nombre"])

                    # medias por lectura (tres sensores)
                    hum_m = float(np.mean(valores[3:6]))   # índices 3,4,5
                    gas_m = float(np.mean(valores[9:12]))  # índices 9,10,11
                    hum_medias.append(hum_m)
                    gas_medias.append(gas_m)

                    # Escribir fila raw en CSV
                    fila = [fase["nombre"], t, pwm] + valores + [hum_m, gas_m]

                    # Actualizar gráficas
                    for i_plot in range(4):
                        for j in range(3):
                            idx_var = orden_graficas[i_plot*3 + j]
                            idx_linea = i_plot*3 + j
                            lineas[idx_linea].set_xdata(datos_t)
                            lineas[idx_linea].set_ydata(datos_vars[idx_var])
                        axs[i_plot].relim()
                        axs[i_plot].autoscale_view()

                    plt.pause(0.01)

                elif linea == "FIN":
                    print(f"✅ Fase '{fase['nombre']}' finalizada")
                    beep()
                    fase_terminada = True
                    break

                elif linea == "STOPPED":
                    print(f"🟥 Fase '{fase['nombre']}' detenida manualmente")
                    fase_terminada = True
                    break

            if not fase_terminada:
                print(f"⚠️ Fase '{fase['nombre']}' terminó inesperadamente")

        except Exception as e:
            print(f"Error durante la fase '{fase['nombre']}': {e}")
            fase_terminada = True

        tiempo_acumulado = datos_t[-1] if datos_t else tiempo_acumulado

        if idx_fase < len(fases) - 1:
            input("Presiona ENTER para continuar a la siguiente fase...")


    # Aplicar suavizado (media móvil) sobre las medias calculadas
    hum_medias = np.array(hum_medias) if len(hum_medias) > 0 else np.array([])
    gas_medias = np.array(gas_medias) if len(gas_medias) > 0 else np.array([])
    hum_suave = media_movil(hum_medias, VENTANA_SUAVIZADO)
    gas_suave = media_movil(gas_medias, VENTANA_SUAVIZADO)

    # Crear CSV final con columnas extra de suavizado
    nombre_csv_final = f"{nombre_ensayo}_{fecha_str}.csv"
    with open(nombre_csv_final, "w", newline="", encoding="utf-8") as ffinal:
        writer = csv.writer(ffinal)
        header_final = ["Fase", "Tiempo(s)", "PWM"] + variables + ["Hum_Media", "Gas_Media", "Hum_Media_Suavizada", "Gas_Media_Suavizada"]
        writer.writerow(header_final)
        n = len(datos_t)
        for i in range(n):
            row = [datos_fase[i], datos_t[i], datos_pwm[i]]
            row += [datos_vars[j][i] for j in range(len(variables))]
            row += [hum_medias[i] if i < len(hum_medias) else "", gas_medias[i] if i < len(gas_medias) else ""]
            row += [float(hum_suave[i]) if i < len(hum_suave) else "", float(gas_suave[i]) if i < len(gas_suave) else ""]
            writer.writerow(row)
    print(f"📁 CSV final con suavizado guardado en: {os.path.abspath(nombre_csv_final)}")

    # --- Cálculos estadísticos y ajustes (para hoja Parámetros) ---
    # Convertir datos_vars a array (12 x N)
    datos_np = np.array(datos_vars) if len(datos_vars) > 0 else np.zeros((len(variables), 0))
    xdata = np.array(datos_t)

    # medias por timestamp (promedio sensores)
    temp_media_ts = np.mean(datos_np[0:3, :], axis=0) if datos_np.shape[1] > 0 else np.array([])
    hum_media_ts = hum_medias  # ya es promedio por timestamp
    pres_media_ts = np.mean(datos_np[6:9, :], axis=0) if datos_np.shape[1] > 0 else np.array([])
    gas_media_ts = gas_medias

    # Ajustes con curve_fit: intentamos ajustar sobre la serie suavizada (si tiene longitud)
    ajustes_activados =(num_fases ==1)
    if ajustes_activados:
        try:
            if len(xdata) > 0 and len(gas_suave) == len(xdata):
                popt_gas, _ = curve_fit(funcion_ajuste, xdata, gas_suave, maxfev=10000)
                y_pred_gas = funcion_ajuste(xdata, *popt_gas)
                r2_gas = calcular_r2(gas_suave, y_pred_gas)
            elif len(xdata) > 0 and len(gas_media_ts) == len(xdata):
                popt_gas, _ = curve_fit(funcion_ajuste, xdata, gas_media_ts, maxfev=10000)
                y_pred_gas = funcion_ajuste(xdata, *popt_gas)
                r2_gas = calcular_r2(gas_media_ts, y_pred_gas)
            else:
                popt_gas = [np.nan, np.nan, np.nan]
                r2_gas = np.nan
        except Exception as e:
            print(f"Error ajuste Gas: {e}")
            popt_gas = [np.nan, np.nan, np.nan]
            r2_gas = np.nan

        try:
            if len(xdata) > 0 and len(hum_suave) == len(xdata):
                popt_hum, _ = curve_fit(funcion_ajuste, xdata, hum_suave, maxfev=10000)
                y_pred_hum = funcion_ajuste(xdata, *popt_hum)
                r2_hum = calcular_r2(hum_suave, y_pred_hum)
            elif len(xdata) > 0 and len(hum_media_ts) == len(xdata):
                popt_hum, _ = curve_fit(funcion_ajuste, xdata, hum_media_ts, maxfev=10000)
                y_pred_hum = funcion_ajuste(xdata, *popt_hum)
                r2_hum = calcular_r2(hum_media_ts, y_pred_hum)
            else:
                popt_hum = [np.nan, np.nan, np.nan]
                r2_hum = np.nan
        except Exception as e:
            print(f"Error ajuste Humedad: {e}")
            popt_hum = [np.nan, np.nan, np.nan]
            r2_hum = np.nan
    else:
        print("⚠️Ensayo con múltiples fases - se omiten los ajustes de Gas y Humedad.")

    # Estadísticas (usar series suavizadas si existen, si no las crudas)
    stats_gas = estadisticas(gas_suave if gas_suave.size > 0 else gas_media_ts)
    stats_hum = estadisticas(hum_suave if hum_suave.size > 0 else hum_media_ts)
    stats_temp = estadisticas(temp_media_ts)
    stats_pres = estadisticas(pres_media_ts)

    # --- Crear Excel con hojas: Parámetros, Datos, Gráficas ---
    nombre_excel = f"{nombre_ensayo}_{fecha_str}.xlsx"
    wb = Workbook()

    # Hoja "Parámetros" 
    ws_param = wb.active
    ws_param.title = "Parámetros"
    ws_param.append(["Nombre ensayo", nombre_ensayo])
    ws_param.append(["Fecha inicio", fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")])
    ws_param.append([])

    # Fases: encabezado y lista
    ws_param.append(["Fase", "Muestra", "Cámara", "Duración (min)", "Ventilador (%)"])
    for f in fases:
        ws_param.append([f["nombre"], f["tipo_muestra"], f["camara_estado"], f["duracion"], f["ventilador"]])
    ws_param.append([])

    # Estadísticas y parámetros de ajuste (encabezado)
    if ajustes_activados:
        ws_param.append(["Variable", "y0", "A", "k", "R²", "Media", "Std", "Var", "Min", "Max"])
        # Gas
        ws_param.append([
            "Gas",
            popt_gas[0], popt_gas[1], popt_gas[2], r2_gas,
            stats_gas["media"], stats_gas["std"], stats_gas["var"], stats_gas["min"], stats_gas["max"]
        ])
        # Humedad
        ws_param.append([
            "Humedad",
            popt_hum[0], popt_hum[1], popt_hum[2], r2_hum,
            stats_hum["media"], stats_hum["std"], stats_hum["var"], stats_hum["min"], stats_hum["max"]
        ])
    
        # Temperatura (sin ajuste)
        ws_param.append([
            "Temperatura",
            "", "", "", "",
            stats_temp["media"], stats_temp["std"], stats_temp["var"], stats_temp["min"], stats_temp["max"]
        ])
        # Presión (sin ajuste)
        ws_param.append([
            "Presion",
            "", "", "", "",
            stats_pres["media"], stats_pres["std"], stats_pres["var"], stats_pres["min"], stats_pres["max"]
        ])
    else:
        ws_param.append(["Variable", "Media", "Std", "Var", "Min", "Max"])
        # Gas
        ws_param.append([
            "Gas",
            stats_gas["media"], stats_gas["std"], stats_gas["var"], stats_gas["min"], stats_gas["max"]
        ])
        # Humedad
        ws_param.append([
            "Humedad",
            stats_hum["media"], stats_hum["std"], stats_hum["var"], stats_hum["min"], stats_hum["max"]
        ])
    
        # Temperatura (sin ajuste)
        ws_param.append([
            "Temperatura",
            stats_temp["media"], stats_temp["std"], stats_temp["var"], stats_temp["min"], stats_temp["max"]
        ])
        # Presión (sin ajuste)
        ws_param.append([
            "Presion",
            stats_pres["media"], stats_pres["std"], stats_pres["var"], stats_pres["min"], stats_pres["max"]
        ])

    # Hoja "Datos"
    ws_datos = wb.create_sheet(title="Datos")
    header_excel = ["Fase", "Tiempo(s)", "PWM"] + variables + ["Hum_Media", "Gas_Media", "Hum_Media_Suavizada", "Gas_Media_Suavizada"]
    ws_datos.append(header_excel)
    unidades = ["", "s", ""] + ["°C","°C","°C","%","%","%","hPa","hPa","hPa","kΩ","kΩ","kΩ"] + ["%","kΩ","%","kΩ"]
    ws_datos.append(unidades)

    N = len(datos_t)
    for i in range(N):
        row = [datos_fase[i], datos_t[i], datos_pwm[i]]
        row += [datos_vars[j][i] for j in range(len(variables))]
        row += [hum_medias[i] if i < len(hum_medias) else "", gas_medias[i] if i < len(gas_medias) else ""]
        row += [float(hum_suave[i]) if i < len(hum_suave) else "", float(gas_suave[i]) if i < len(gas_suave) else ""]
        ws_datos.append(row)

    # Hoja "Gráficas"
    fig_final, axs_final = plt.subplots(4, 1, figsize=(12,12), sharex=True)
    fig_final.suptitle(f"Ensayo: {nombre_ensayo}")

    #----Gas----
    axs_final[0].plot(datos_t, datos_np[9], 'b-', alpha=0.8, label='Gas_A raw')
    axs_final[0].plot(datos_t, datos_np[10], 'r-', alpha=0.8, label='Gas_B raw')
    axs_final[0].plot(datos_t, datos_np[11], 'g-', alpha=0.8, label='Gas_C raw')
    axs_final[0].plot(datos_t, gas_suave, 'y-', linewidth=2, label='Gas Media Suavizada')
    if ajustes_activados and len(xdata) > 0 and not np.isnan(popt_gas).all():
        axs_final[0].plot(xdata, y_pred_gas, 'p--', linewidth=2, label='Ajuste Gas')
    axs_final[0].set_ylabel("Gas (kΩ)")
    axs_final[0].legend()
    axs_final[0].grid(True)
    
    
    # --- Humedad ---
    axs_final[1].plot(datos_t, datos_np[3], 'b-', alpha=0.8, label='Hum_A raw')
    axs_final[1].plot(datos_t, datos_np[4], 'r-', alpha=0.8, label='Hum_B raw')
    axs_final[1].plot(datos_t, datos_np[5], 'g-', alpha=0.8, label='Hum_C raw')
    axs_final[1].plot(datos_t, hum_suave, 'y-', linewidth=2, label='Hum Media Suavizada')
    if ajustes_activados and len(xdata) > 0 and not np.isnan(popt_hum).all():
        axs_final[1].plot(xdata, y_pred_hum, 'p--', linewidth=2, label='Ajuste Hum')
    axs_final[1].set_ylabel("Humedad (%)")
    axs_final[1].legend()
    axs_final[1].grid(True)

    # --- Temperatura ---
    axs_final[2].plot(datos_t, datos_np[0], 'b-', alpha=0.8, label='Temp_A raw')
    axs_final[2].plot(datos_t, datos_np[1], 'r-', alpha=0.8, label='Temp_B raw')
    axs_final[2].plot(datos_t, datos_np[2], 'g-', alpha=0.8, label='Temp_C raw')
    axs_final[2].plot(datos_t, np.mean(datos_np[0:3,:], axis=0), 'y-', linewidth=2, label='Temp Media')
    axs_final[2].set_ylabel("Temperatura (°C)")
    axs_final[2].legend()
    axs_final[2].grid(True)

    # --- Presión ---
    axs_final[3].plot(datos_t, datos_np[6], 'b-', alpha=0.8, label='Pres_A raw')
    axs_final[3].plot(datos_t, datos_np[7], 'r-', alpha=0.8, label='Pres_B raw')
    axs_final[3].plot(datos_t, datos_np[8], 'g-', alpha=0.8, label='Pres_C raw')
    axs_final[3].plot(datos_t, np.mean(datos_np[6:9,:], axis=0), 'y-', linewidth=2, label='Pres Media')
    axs_final[3].set_ylabel("Presión (hPa)")
    axs_final[3].set_xlabel("Tiempo (s)")
    axs_final[3].legend()
    axs_final[3].grid(True)

    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    img_path = f"grafica_{fecha_str}.png"
    fig_final.savefig(img_path, bbox_inches="tight")
    plt.close(fig_final)  # cerrar figura para liberar memoria

    # Insertar en Excel
    ws_graf = wb.create_sheet(title="Gráficas")
    try:
        img = ExcelImage(img_path)
        img.anchor = 'A1'
        ws_graf.add_image(img)
    except Exception as e:
        print(f"⚠️ No se pudo insertar imagen en Excel: {e}")

    # Ordenar hojas
        wb._sheets = [ws_param, ws_datos, ws_graf]

    # Guardar workbook 
    try:
        wb.save(nombre_excel)
        print(f"📊 Excel guardado correctamente: {os.path.abspath(nombre_excel)}")
    except Exception as e:
        print(f"❌ Error al guardar Excel: {e}")
    # borrar png temporal si existe
    try:
        if os.path.exists(img_path):
            os.remove(img_path)
    except Exception:
        pass

    beep()
    plt.ioff()
    print("✅ Ensayo finalizado correctamente.")

def main():
    ser = None
    while True:
        try:
            if ser is None or not ser.is_open:
                ser = serial.Serial(PUERTO, BAUDIOS, timeout=TIMEOUT)
                time.sleep(2)
                ser.reset_input_buffer()
                esperar_ready(ser)
            ejecutar_ensayo(ser)
            print("\nEnsayo finalizado. Puedes iniciar otro ensayo o presionar Ctrl+C para salir.\n")
        except serial.SerialException as e:
            print(f"Error de conexión: {e}")
            print("Reconectando en 5 segundos...")
            if ser:
                try: ser.close()
                except: pass
            ser = None
            time.sleep(5)
        except KeyboardInterrupt:
            print("Programa terminado por usuario.")
            if ser:
                ser.close()
            break
        except Exception as e:
            print(f"Error inesperado: {e}")
            if ser:
                try: ser.close()
                except: pass
            ser = None
            time.sleep(5)

if __name__ == "__main__":
    main()
